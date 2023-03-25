import os
import sys
import cv2
from simple_facerec import SimpleFacerec
from openpyxl import Workbook, load_workbook
from datetime import datetime, date, timedelta

# GUI
from interfata import Ui_MainWindow

# PyQt5 imports
from PyQt5.QtCore import *
from PyQt5.QtGui import QPixmap, QImage
from PyQt5.QtWidgets import QMainWindow, QFileDialog, QApplication


class Window(QMainWindow, Ui_MainWindow):
	def __init__(self, parent=None):
		super(Window, self).__init__(parent)

		self.setupUi(self)

		# Buttons
		self.UploadPhotoCB.clicked.connect(self.photo)
		self.UploadVideoCB.clicked.connect(self.video)
		self.RecognitionPhotoCB.clicked.connect(self.detectionPhoto)
		self.RecognitionVideoCB.clicked.connect(self.detectionVideo)
		self.StartCameraCB.clicked.connect(self.detectionLive)

		self.filename = None

		# thread control
		self.threadpool = QThreadPool()

		# Thread workers
		self.runnerP = None
		self.runnerV = None
		self.runnerC = None

	def photo(self):
		# incarcare poza
		self.filename, _ = QFileDialog.getOpenFileName(
			self,
			"Open File",
			os.getcwd(),
			"All Files (*)"
		)
		# Open the image
		pixmap = QPixmap(self.filename)
		# Add pic to label
		self.WhereIsPhoto.setPixmap(pixmap)

	def video(self):
		# incarcare video
		self.filename, _ = QFileDialog.getOpenFileName(
			self,
			"Open File",
			"C:/Users/Denis/Desktop/Proiect/Imagini",
			"All Files (*)"
		)

	def updateImageLabel(self, frame):
		pixmapImg = self.convertCV2PIXMAP(frame)
		self.WhereIsPhoto.setPixmap(QPixmap.fromImage(pixmapImg))

	def updateVideoLabel(self, frame):
		pixmapImg = self.convertCV2PIXMAP(frame)
		self.WhereIsVideo.setPixmap(QPixmap.fromImage(pixmapImg))

	def updateCameraLabel(self, frame):
		pixmapImg = self.convertCV2PIXMAP(frame)
		self.WhereIsCamera.setPixmap(QPixmap.fromImage(pixmapImg))

	@staticmethod
	def convertCV2PIXMAP(frame):
		# from openCv to pixmap
		rgb_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
		h, w, ch = rgb_image.shape
		return QImage(rgb_image.data, w, h, ch * w, QImage.Format_RGB888)

	def detectionPhoto(self):
		if self.filename is not None:
			self.runnerP = Processing(isPhoto=True, path=self.filename)
			self.threadpool.start(self.runnerP)
			self.filename = None

	def detectionVideo(self):
		if self.filename is not None:
			# Prelucrare pe alt fir de executie pentru evitarea blocajului
			self.runnerV = Processing(isPhoto=False, path=self.filename)
			self.threadpool.start(self.runnerV)
			self.filename = None

			# Oprire fir de executie
			self.StopVideoCB.pressed.connect(self.runnerV.kill)

	def detectionLive(self):
		# Prelucrare pe alt fir de executie pentru evitarea blocajului
		self.runnerC = Processing(isPhoto=False, isCam=True)
		self.threadpool.start(self.runnerC)

		# Oprire fir de executie
		self.StopCameraCB.pressed.connect(self.runnerC.kill)


# Clasa thread pentru procesarea informatiilor (evitarea blocarii GUI-ului)
class Processing(QRunnable):
	finished = pyqtSignal()

	def __init__(self, isPhoto, isCam=False, path=""):
		super().__init__()
		self.isPhoto = isPhoto  	# process a photo or a video?
		self.isCam = isCam  		# process a local video or connect to a live feed?
		if (isPhoto or not isCam) and path == "":
			print("Path ul nu poate fi gol!")
			exit()
		self.path = path 			# path to the local photo/video
		self.is_killed = False
		self.cap = None				# reads frames from video or live feed
		self.process_this_frame = True  # for skiping frames

		self.startTime = None
		self.endTime = None

	@pyqtSlot()
	def run(self):
		if self.isPhoto:
			self.photoProcessing()
		else:
			if self.isCam:
				self.cap = cv2.VideoCapture(1)
				if not self.cap.isOpened():
					print("Cannot open camera")
					exit()
				self.videoProcessing()

			else:
				self.cap = cv2.VideoCapture(self.path)
				self.videoProcessing()

			# When everything done, release the capture
			self.cap.release()
			cv2.destroyAllWindows()

	def photoProcessing(self):
		frame = cv2.imread(self.path)
		frame, _ = self.visualize(frame)
		windowVar.updateImageLabel(frame)

	def videoProcessing(self):
		while True:
			# Time at the start of processing
			if self.isCam:
				self.startTime = datetime.now()
			# Capture frame-by-frame
			ret, frame = self.cap.read()
			# if frame is read correctly ret is True
			if not ret:
				break
			tm.reset()
			tm.start()
			frame, names = self.visualize(frame)
			tm.stop()
			cv2.putText(frame, 'FPS: {:.2f}'.format(tm.getFPS()), (1, 16), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)
			if self.isCam:
				windowVar.updateCameraLabel(frame)
			else:
				windowVar.updateVideoLabel(frame)

			# Time at the end of processing
			if self.isCam:
				self.endTime = datetime.now()
				for name in list(set(names)):
					if name == 'Unknown':
						continue
					# Check if name exists
					index = 0
					for i in range(2, activeSheet.max_row + 1):
						if activeSheet.cell(row=i, column=1).value == name:
							index = i
							break

					if index == 0:
						totalSec = (self.endTime - self.startTime).total_seconds()
						isPresent = True if totalSec > 60 else False
						activeSheet.append(
							[
								name,
								str(isPresent),
								str(timedelta(seconds=totalSec)),  # calculez din secunde in h,min,s
								totalSec
							]
						)
					else:
						activeSheet[index][3].value = activeSheet[index][3].value + (self.endTime - self.startTime).total_seconds()
						activeSheet[index][2].value = str(timedelta(seconds=activeSheet[index][3].value))
						activeSheet[index][1].value = str(True) if activeSheet[index][3].value > 60 else str(False)

				wb.save('data.xlsx')

	@staticmethod
	def visualize(frame):
		# Detect Faces
		face_locations, face_names = sfr.detect_known_faces(frame)
		for face_loc, name in zip(face_locations, face_names):
			y1, x2, y2, x1 = face_loc[0], face_loc[1], face_loc[2], face_loc[3]

			cv2.putText(frame, name, (x1, y1 - 10), cv2.FONT_HERSHEY_DUPLEX, 1, (0, 0, 200), 2)
			cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 0, 200), 4)
		return frame, face_names

	def kill(self):
		self.is_killed = True
		self.cap.release()
		cv2.destroyAllWindows()


def quitFunction():
	try:
		windowVar.runnerC.stop()
		windowVar.runnerV.stop()
	except (Exception,):
		pass
	finally:
		wb.save(filename="data.xlsx")
		wb.close()


if __name__ == "__main__":
	# Simple Face Recognition class
	sfr = SimpleFacerec()
	#incarca pozele persoanelor cunoscute
	sfr.load_encoding_images("Faces/")

	# Initializare excel
	if not os.path.isfile('data.xlsx'):
		wb = Workbook()
		wb.save(filename='data.xlsx')
	else:
		wb = load_workbook(filename='data.xlsx')
	today = date.today().strftime("%d-%m-%Y")
	if today not in wb.sheetnames:
		wb.create_sheet(title=today)
		activeSheet = wb[today]
		activeSheet.append(['Nume', 'Prezent?', 'Timp', 'Total secunde'])
		wb.save('data.xlsx')
	else:
		activeSheet = wb[today]

	tm = cv2.TickMeter()

	# Main QT Application
	app = QApplication(sys.argv)
	app.aboutToQuit.connect(quitFunction)
	windowVar = Window()
	windowVar.show()
	app.exec_()
